import csv
import json
import os
from datetime import datetime

import openpyxl

# Константы для статусов
STATUSES = ["EXECUTED", "CANCELED", "PENDING"]
DATA_DIR = "../data"


def load_json_transactions(file_path):
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            transactions = json.load(file)
            return transactions
    except FileNotFoundError:
        print(f"Ошибка: файл {file_path} не найден")
        return []
    except json.JSONDecodeError:
        print("Ошибка: файл поврежден или не является JSON")
        return []


def load_csv_transactions(file_path):
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            reader = csv.DictReader(file)
            transactions = [row for row in reader]
            return transactions
    except FileNotFoundError:
        print(f"Ошибка: файл {file_path} не найден")
        return []
    except Exception as e:
        print(f"Ошибка при чтении CSV: {str(e)}")
        return []


def load_xlsx_transactions(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        transactions = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            transactions.append(dict(zip(headers, row)))

        return transactions
    except FileNotFoundError:
        print(f"Ошибка: файл {file_path} не найден")
        return []
    except Exception as e:
        print(f"Ошибка при чтении XLSX: {str(e)}")
        return []


def filter_transactions(transactions, status, currency=None, search_text=None):
    filtered = []
    for tx in transactions:
        # Проверяем, что поле 'status' не равно None
        if tx.get("status") is None or tx.get("status").upper() != status.upper():
            continue

        if currency and tx.get("currency") != currency:
            continue

        if search_text and search_text.lower() not in tx.get("description", "").lower():
            continue

        filtered.append(tx)
    return filtered


def sort_transactions(transactions, ascending=True):
    try:
        return sorted(transactions, key=lambda x: datetime.strptime(x["date"], "%d.%m.%Y"), reverse=not ascending)
    except Exception as e:
        print(f"Ошибка при сортировке: {str(e)}")
        return transactions


def print_transactions(transactions):
    if not transactions:
        print("Программа: Не найдено ни одной транзакции, подходящей под ваши условия фильтрации")
        return

    print("\nПрограмма: Распечатываю итоговый список транзакций...")
    print(f"Программа: Всего банковских операций в выборке: {len(transactions)}\n")
    for tx in transactions:
        date = tx.get("date")
        description = tx.get("description")
        amount = tx.get("amount")
        currency = tx.get("currency")
        print(f"{date} {description}\nСчет **4321\nСумма: {amount} {currency}\n")


def main():
    print("Программа: Привет! Добро пожаловать в программу работы с банковскими транзакциями.")
    print("Программа: Выберите необходимый пункт меню:")
    print("1. Получить информацию о транзакциях из JSON-файла")
    print("2. Получить информацию о транзакциях из CSV-файла")
    print("3. Получить информацию о транзакциях из XLSX-файла")

    while True:
        try:
            choice = int(input("Пользователь: "))
            if choice in [1, 2, 3]:
                break
            else:
                print("Программа: Пожалуйста, выберите один из предложенных пунктов меню.")
        except ValueError:
            print("Программа: Пожалуйста, введите число (1, 2 или 3).")

    if choice == 1:
        print("Программа: Для обработки выбран JSON-файл.")
        file_extension = ".json"
    elif choice == 2:
        print("Программа: Для обработки выбран CSV-файл.")
        file_extension = ".csv"
    else:
        print("Программа: Для обработки выбран XLSX-файл.")
        file_extension = ".xlsx"

    while True:
        status = input(
            f"Программа: Введите статус, по которому необходимо выполнить фильтрацию.\nДоступные для фильтровки статусы: {', '.join(STATUSES)}\nПользователь: "
        )
        if status.upper() in STATUSES:
            print(f'Программа: Операции отфильтрованы по статусу "{status.upper()}"')
            break
        else:
            print(f'Программа: Статус операции "{status}" недоступен.')

    # Загрузка транзакций
    file_name = input("Программа: Введите имя файла: ")
    file_path = os.path.join(DATA_DIR, file_name + file_extension)

    if file_extension == ".json":
        transactions = load_json_transactions(file_path)
    elif file_extension == ".csv":
        transactions = load_csv_transactions(file_path)
    else:
        transactions = load_xlsx_transactions(file_path)

    if not transactions:
        return

    # Фильтрация по статусу
    filtered_transactions = filter_transactions(transactions, status)

    # Сортировка по дате
    sort_option = input("Программа: Отсортировать операции по дате? Да/Нет\nПользователь: ").lower()
    if sort_option == "да":
        order_option = input("Программа: Отсортировать по возрастанию или по убыванию?\nПользователь: ").lower()
        ascending = order_option == "по возрастанию"
        filtered_transactions = sort_transactions(filtered_transactions, ascending)

    # Фильтрация по валюте
    currency_option = input("Программа: Выводить только рублевые транзакции? Да/Нет\nПользователь: ").lower()
    if currency_option == "да":
        filtered_transactions = filter_transactions(filtered_transactions, status, currency="руб")

    # Фильтрация по тексту в описании
    search_option = input(
        "Программа: Отфильтровать список транзакций по определенному слову в описании? Да/Нет\nПользователь: "
    ).lower()
    if search_option == "да":
        search_text = input("Программа: Введите слово для поиска:\nПользователь: ")
        filtered_transactions = filter_transactions(filtered_transactions, status, search_text=search_text)

    # Вывод транзакций
    print_transactions(filtered_transactions)


if __name__ == "__main__":
    # Создаем директорию "data", если она отсутствует
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)
    main()